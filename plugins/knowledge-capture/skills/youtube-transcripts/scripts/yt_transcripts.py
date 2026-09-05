#!/usr/bin/env python3
"""Stáhne přehled + přepisy videí z YouTube playlistu, kanálu nebo jednoho videa.

Použití:
    python3 yt_transcripts.py <url> --out <složka> [--list-only] [--limit N]
                              [--lang en,cs] [--delay 5] [--name "Název"]

<url>  playlist (…list=…), kanál (youtube.com/@handle nebo …/videos), nebo jedno video.

Výstup ve složce --out:
    00-prehled.xlsx   všechna videa (titulek, odkaz, datum, délka, zhlédnutí, dostupnost)
    00-index.md       stejný přehled v markdownu s odkazy na přepisy
    NN-slug.md        jeden soubor na video: frontmatter, popis, přepis s časovými značkami

Skript je idempotentní: už stažené přepisy přeskočí, takže po přerušení nebo
blokaci ze strany YouTube stačí spustit znovu se stejnými parametry.

Závislosti si sám nainstaluje do ~/.cache/datawizard/youtube-transcripts-venv, preferuje Homebrew
python3.12+ (systémový 3.9 už yt-dlp nepodporuje).
"""
import argparse
import html
import json
import os
import re
import shutil
import subprocess
import sys
import time
import unicodedata
from datetime import date
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from urllib.request import Request, urlopen

SKILL_DIR = Path(__file__).resolve().parent.parent
VENV = Path.home() / ".cache" / "datawizard" / "youtube-transcripts-venv"
DEPS = ["yt-dlp", "youtube-transcript-api", "openpyxl"]
UA = {"User-Agent": "Mozilla/5.0", "Accept-Language": "en"}


# --------------------------------------------------------------------------- venv
def ensure_venv():
    """Vytvoří venv s novým Pythonem a znovu spustí skript uvnitř něj."""
    if sys.prefix == str(VENV):
        return
    py = VENV / "bin" / "python3"
    if not py.exists():
        base = next((p for p in ("python3.13", "python3.12", "python3.11", "python3.10")
                     if shutil.which(p)), "python3")
        print(f"[setup] zakládám venv ({base}) v {VENV}", flush=True)
        subprocess.run([base, "-m", "venv", str(VENV)], check=True)
        subprocess.run([str(py), "-m", "pip", "install", "-q", "--upgrade", "pip", *DEPS],
                       check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    os.execv(str(py), [str(py), __file__, *sys.argv[1:]])


ensure_venv()

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from youtube_transcript_api import YouTubeTranscriptApi  # noqa: E402


# --------------------------------------------------------------------------- helpers
def slug(s, n=60):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s[:n].rstrip("-") or "video"


def fmt_dur(secs):
    secs = int(secs or 0)
    h, m, s = secs // 3600, (secs % 3600) // 60, secs % 60
    return f"{h}:{m:02d}:{s:02d}" if h else f"{m}:{s:02d}"


def video_ids(url, limit):
    """Vrátí seznam ID videí. Jedno video → [id]; playlist/kanál → přes yt-dlp."""
    q = parse_qs(urlparse(url).query)
    if "list" not in q and ("v" in q or "youtu.be" in url):
        return [q["v"][0] if "v" in q else urlparse(url).path.strip("/")]
    if "list" in q:
        url = f"https://www.youtube.com/playlist?list={q['list'][0]}"
    cmd = [str(VENV / "bin" / "yt-dlp"), "--flat-playlist", "--print", "%(id)s", url]
    if limit:
        cmd += ["--playlist-end", str(limit)]
    out = subprocess.run(cmd, capture_output=True, text=True)
    ids = [x for x in out.stdout.split() if re.fullmatch(r"[\w-]{11}", x)]
    if not ids:
        sys.exit(f"Nenašel jsem žádná videa. yt-dlp: {out.stderr.strip()[-500:]}")
    return ids


def fetch_meta(vid):
    """Metadata z watch stránky (ytInitialPlayerResponse). Bez API klíče, bez přihlášení."""
    page = urlopen(Request(f"https://www.youtube.com/watch?v={vid}", headers=UA),
                   timeout=30).read().decode("utf-8", "ignore")
    m = re.search(r"ytInitialPlayerResponse\s*=\s*(\{.+?\})\s*;\s*(?:var\s|</script>)", page, re.S)
    data = json.loads(m.group(1)) if m else {}
    vd = data.get("videoDetails", {})
    mf = data.get("microformat", {}).get("playerMicroformatRenderer", {})
    ps = data.get("playabilityStatus", {})
    status = ps.get("status", "UNKNOWN")
    reason = ps.get("reason", "")
    return {
        "id": vid,
        "url": f"https://www.youtube.com/watch?v={vid}",
        "title": html.unescape(vd.get("title", "")),
        "channel": vd.get("author", ""),
        "date": (mf.get("publishDate") or "")[:10],
        "secs": int(vd.get("lengthSeconds") or 0),
        "views": int(vd.get("viewCount") or 0),
        "desc": vd.get("shortDescription", ""),
        "ok": status == "OK",
        "status": status,
        "reason": reason,
    }


def fetch_transcript(api, vid, langs, retries=6):
    """Přepis v preferovaném jazyce, jinak jakýkoli dostupný. Backoff při blokaci IP."""
    for attempt in range(retries):
        try:
            try:
                t = api.fetch(vid, languages=langs)
            except Exception as e:
                if "IpBlocked" in type(e).__name__ or "RequestBlocked" in type(e).__name__:
                    raise
                # jazyk nedostupný → vezmi cokoli
                tl = api.list(vid)
                t = next(iter(tl)).fetch()
            return t
        except Exception as e:
            name = type(e).__name__
            if name in ("IpBlocked", "RequestBlocked", "ConnectionError", "YouTubeRequestFailed"):
                wait = 45 * (attempt + 1)
                print(f"    blokace ({name}), čekám {wait}s…", flush=True)
                time.sleep(wait)
                continue
            if name in ("TranscriptsDisabled", "NoTranscriptFound", "VideoUnavailable"):
                return None
            raise
    raise RuntimeError("blokace trvá i po opakovaných pokusech")


def transcript_to_md(t, chunk=60):
    """Segmenty → odstavce po ~chunk sekundách s časovou značkou."""
    paras, cur, start = [], [], 0.0
    for s in t:
        if s.start - start >= chunk and cur:
            paras.append((start, " ".join(cur)))
            cur, start = [], s.start
        cur.append(s.text.replace("\n", " ").strip())
    if cur:
        paras.append((start, " ".join(cur)))
    body = "\n\n".join(f"**[{fmt_dur(st)}]** {tx}" for st, tx in paras)
    words = sum(len(tx.split()) for _, tx in paras)
    return body, words


def write_video_md(path, n, m, t, body, words):
    lang = f"{t.language} ({'auto' if t.is_generated else 'ruční'})"
    title = m["title"].replace('"', "'")
    path.write_text(f"""---
title: "{title}"
date: {m['date']}
type: research
source: youtube
channel: "{m['channel']}"
video_id: {m['id']}
url: {m['url']}
published: {m['date']}
duration: "{fmt_dur(m['secs'])}"
views: {m['views']}
transcript: {lang}
words: {words}
---

# {n}. {m['title']}

- Odkaz: {m['url']}
- Kanál: {m['channel']}
- Publikováno: {m['date']}, délka {fmt_dur(m['secs'])}, zhlédnutí {m['views']}

## Popis videa

{m['desc'].strip() or '(bez popisu)'}

## Přepis ({lang})

{body}
""")


def write_xlsx(path, rows, name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Videa"
    ws.append(["#", "Titulek", "Odkaz", "Kanál", "Datum vydání", "Délka", "Zhlédnutí",
               "Dostupné", "Přepis", "Poznámka"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
    for r in rows:
        ws.append([r["n"], r["title"], r["url"], r["channel"], r["date"], r["dur"], r["views"],
                   "ano" if r["ok"] else "ne", r["file"] or "", r["note"]])
        cell = ws.cell(row=ws.max_row, column=3)
        cell.hyperlink = r["url"]
        cell.style = "Hyperlink"
        if not r["ok"]:
            for c in ws[ws.max_row]:
                c.fill = PatternFill("solid", fgColor="FCE4D6")
    for i, w in enumerate([5, 60, 44, 22, 13, 9, 11, 10, 40, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def write_index(path, rows, name, src):
    lines = [f"---\ntitle: {name} – index videí\ndate: {date.today()}\ntype: research\n---\n",
             f"# {name} – index videí\n", f"Zdroj: {src}  ", "Excel: [00-prehled.xlsx](00-prehled.xlsx)\n",
             "| # | Titulek | Kanál | Publikováno | Délka | Slov | Přepis |",
             "|---|---|---|---|---|---|---|"]
    for r in rows:
        link = f"[{r['file']}]({r['file']})" if r["file"] else r["note"]
        lines.append(f"| {r['n']} | [{r['title']}]({r['url']}) | {r['channel']} | {r['date']} | "
                     f"{r['dur']} | {r['words'] or ''} | {link} |")
    path.write_text("\n".join(lines) + "\n")


# --------------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("url")
    ap.add_argument("--out", required=True, help="výstupní složka (vytvoří se)")
    ap.add_argument("--name", help="název sady pro index (default: z URL)")
    ap.add_argument("--list-only", action="store_true", help="jen přehled (xlsx + index), bez přepisů")
    ap.add_argument("--limit", type=int, default=0, help="zpracovat jen prvních N videí")
    ap.add_argument("--lang", default="en,cs", help="preferované jazyky přepisu, čárkou")
    ap.add_argument("--delay", type=float, default=5, help="pauza mezi přepisy v s (limity YouTube)")
    a = ap.parse_args()

    out = Path(a.out)
    out.mkdir(parents=True, exist_ok=True)
    langs = [x.strip() for x in a.lang.split(",") if x.strip()]
    api = YouTubeTranscriptApi()

    ids = video_ids(a.url, a.limit)
    print(f"Videí: {len(ids)}", flush=True)
    rows, blocked = [], 0
    for n, vid in enumerate(ids, 1):
        m = fetch_meta(vid)
        row = {"n": n, "title": m["title"] or "(bez titulku)", "url": m["url"], "channel": m["channel"],
               "date": m["date"], "dur": fmt_dur(m["secs"]) if m["secs"] else "", "views": m["views"] or None,
               "ok": m["ok"], "file": None, "words": None, "note": ""}
        if not m["ok"]:
            row["note"] = "nestaženo – " + ({"LOGIN_REQUIRED": "soukromé video (vyžaduje přístup od majitele)",
                                             "ERROR": "video nedostupné / smazané"}.get(m["status"], m["reason"] or m["status"]))
            print(f"{n:3d} {vid}  --  {row['note']}", flush=True)
            rows.append(row)
            continue

        existing = list(out.glob(f"{n:02d}-*.md"))
        if existing:
            row["file"] = existing[0].name
            w = re.search(r"^words: (\d+)", existing[0].read_text(), re.M)
            row["words"] = int(w.group(1)) if w else None
            print(f"{n:3d} {vid}  ok  (už staženo) {m['title'][:60]}", flush=True)
        elif a.list_only:
            print(f"{n:3d} {vid}  ok  {m['date']}  {m['title'][:60]}", flush=True)
        else:
            try:
                t = fetch_transcript(api, vid, langs)
            except RuntimeError as e:
                row["note"] = f"přepis nestažen – {e}; spusť znovu"
                blocked += 1
                print(f"{n:3d} {vid}  !!  {row['note']}", flush=True)
                rows.append(row)
                continue
            if t is None:
                row["note"] = "video nemá titulky ani auto-přepis"
                print(f"{n:3d} {vid}  --  {row['note']}", flush=True)
            else:
                body, words = transcript_to_md(t)
                fname = f"{n:02d}-{slug(m['title'])}.md"
                write_video_md(out / fname, n, m, t, body, words)
                row["file"], row["words"] = fname, words
                print(f"{n:3d} {vid}  ok  {m['date']}  {words:5d} slov  {m['title'][:50]}", flush=True)
            time.sleep(a.delay)
        rows.append(row)

    name = a.name or (rows[0]["channel"] if rows and rows[0]["channel"] else "YouTube")
    write_xlsx(out / "00-prehled.xlsx", rows, name)
    write_index(out / "00-index.md", rows, name, a.url)

    got = sum(1 for r in rows if r["file"])
    unavailable = sum(1 for r in rows if not r["ok"])
    print(f"\nHotovo: {len(rows)} videí, přepisů {got}, nedostupných videí {unavailable}, "
          f"blokováno {blocked}. Výstup: {out}")
    if blocked:
        print("YouTube blokoval část požadavků. Spusť stejný příkaz znovu, stažené se přeskočí.")


if __name__ == "__main__":
    main()
